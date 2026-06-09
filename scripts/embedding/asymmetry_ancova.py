"""鏡射不對稱的 ANCOVA（年齡共變量）分析：表、散點、跨模型／ArcFace 總圖。

平行斜率 score = β0 + β1·age + β2·group，β2＝年齡校正後組別差；另以交互模型取 β3＝兩組年齡
斜率差（H0:β3=0，與平行模型分開擬合、不動 β0/β1/β2）。底層 import 自 asymmetry_stat。

--mode 輸出（皆在 <feature_stat>/<cohort>/<bg>/…）：
  table     <model>/asymmetry_score_stat_ancova.xlsx + ancova/<method>.png（AD-vs-HC 散點）
  age-slope <model>/asymmetry_age_slope.{xlsx,png}（各組 score~age，β1 單尾）
  grid      ancova_grid.png（4 模型 × 8 method，AD-vs-HC）
  arcface   arcface/ancova/ancova_grid_arcface{,_full,_matched}.png（32 格 + full/matched 切分）

Usage:
    python scripts/embedding/asymmetry_ancova.py --mode {table|age-slope|grid|arcface} [--model arcface]
"""

import argparse
import logging
import sys
from collections import defaultdict
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import numpy as np
from openpyxl import Workbook
from scipy import stats as sp_stats

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))  # scripts/
from _paths import PROJECT_ROOT  # noqa: F401

from src.config import (
    EMBEDDING_FEATURE_STAT_DIR, cohort_path,
    P_VISIT_TOKENS, P_SCORE_TOKENS, HC_VISIT_TOKENS, HC_SCORE_TOKENS,
)
from src.common.cohort import cohort_list, base_id_of, group_of
from asymmetry_stat import (
    DEFAULT_TOKENS, METHODS, MODEL_DISPLAY,
    scorer_scores, full_comparisons, matched_comparisons, _pstr,
    _BORDER, _CENTER, _F_HEADER, _F_BIG, _F_DATA, _ROW_H_HEADER, _ROW_H_DATA,
)

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

MODELS = ["arcface", "dlib", "topofr", "vggface"]      # grid 列順序（上→下）
COL_LABELS = [label for label, *_ in METHODS]          # grid 欄＝8 個 method 顯示名

# arcface 32 格：欄＝4 norm × 2 asymmetry vector；case=紅(group=1) / control=藍(group=0)
_ARCFACE = "arcface"
_CASE, _CTRL = "#C44E52", "#4C72B0"
ARCFACE_COLS = [
    ("L1 · diff", "differences", "l1_norm"),
    ("L2 · diff", "differences", "l2_norm"),
    ("L1 · rel_diff", "relative_differences", "l1_norm"),
    ("L2 · rel_diff", "relative_differences", "l2_norm"),
]
ARCFACE_DIFF_COLS = ARCFACE_COLS[:2]   # 切分圖只用 L1·diff、L2·diff

# ── ANCOVA 擬合 ────────────────────────────────────────────────────────────

def per_subject_age(cohort_df) -> dict:
    """每 subject 年齡 = 其各 cohort visit 的 Age 平均。"""
    by = defaultdict(list)
    for sid, age in zip(cohort_df["ID"], cohort_df["Age"]):
        by[base_id_of(sid)].append(float(age))
    return {b: float(np.mean(v)) for b, v in by.items()}


def _ancova_fit(scores, age, s1, s2):
    """OLS score ~ 1 + age + group(arm1=1) 於全體 subject（平行斜率、無交互作用）。

    Returns:
        (n, unadjΔ, unadj_p, b0, b1_age, b2_group, group_p, age_p)；資料不足回 None。
        unadjΔ/unadj_p＝未校正 arm1−arm2 均差與 Welch t p；b2/group_p＝年齡校正後的組別效應。
    """
    a1 = [b for b in s1 if b in scores and b in age]
    a2 = [b for b in s2 if b in scores and b in age]
    if len(a1) < 2 or len(a2) < 2:
        return None
    y = np.array([scores[b] for b in a1 + a2])
    g = np.array([1.0] * len(a1) + [0.0] * len(a2))
    ag = np.array([age[b] for b in a1 + a2])
    X = np.column_stack([np.ones_like(y), ag, g])
    beta, *_ = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta
    dof = len(y) - X.shape[1]
    se = np.sqrt(np.diag((resid @ resid) / dof * np.linalg.inv(X.T @ X)))
    p = 2 * sp_stats.t.sf(np.abs(beta / se), dof)
    unadj = y[g == 1].mean() - y[g == 0].mean()
    unadj_p = sp_stats.ttest_ind(y[g == 1], y[g == 0], equal_var=False).pvalue
    return len(y), unadj, unadj_p, beta[0], beta[1], beta[2], p[2], p[1]


def _coef(b):
    return f"{b:+.4f}" if abs(b) >= 1e-3 else f"{b:+.2e}"


def _one_sided_p(b, p2):
    """雙尾 p2 → 單尾 p（H1: 係數 > 0）：b≥0 取 p2/2，否則 1 − p2/2。β2、β3 通用。"""
    return p2 / 2 if b >= 0 else 1 - p2 / 2


def _slope_diff_fit(scores, age, s1, s2):
    """交互模型 score ~ 1 + age + group + age×group，回 (b3, b3_p2)。

    b3＝age×group 係數＝兩組年齡斜率差，b3_p2 雙尾。與 _ancova_fit 分開擬合、不影響其
    β0/β1/β2。資料不足（任一組 <2 或自由度 ≤0）回 None。
    """
    a1 = [b for b in s1 if b in scores and b in age]
    a2 = [b for b in s2 if b in scores and b in age]
    if len(a1) < 2 or len(a2) < 2:
        return None
    y = np.array([scores[b] for b in a1 + a2])
    g = np.array([1.0] * len(a1) + [0.0] * len(a2))
    ag = np.array([age[b] for b in a1 + a2])
    X = np.column_stack([np.ones_like(y), ag, g, ag * g])
    dof = len(y) - X.shape[1]
    if dof <= 0:
        return None
    beta, *_ = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta
    se = np.sqrt(np.diag((resid @ resid) / dof * np.linalg.inv(X.T @ X)))
    b3 = beta[3]
    b3_p2 = 2 * sp_stats.t.sf(np.abs(b3 / se[3]), dof)
    return b3, b3_p2


def _slope_fit(scores, age, members):
    """單組簡單迴歸 score ~ 1 + age，回 (n, b0, b1, b1_p2)（b1_p2 雙尾）。資料不足（<3）回 None。"""
    ids = [b for b in members if b in scores and b in age]
    if len(ids) < 3:
        return None
    y = np.array([scores[b] for b in ids])
    ag = np.array([age[b] for b in ids])
    X = np.column_stack([np.ones_like(y), ag])
    dof = len(y) - 2
    if dof <= 0:
        return None
    beta, *_ = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta
    se = np.sqrt(np.diag((resid @ resid) / dof * np.linalg.inv(X.T @ X)))
    b1_p2 = 2 * sp_stats.t.sf(np.abs(beta[1] / se[1]), dof)
    return len(y), beta[0], beta[1], b1_p2


_GROUP_COLOR = {"AD": "#C44E52", "NAD": "#55A868", "ACS": "#4C72B0", "HC": "#333333"}


def _age_slope_groups(age):
    """per-group age-slope 的 4 組（互斥的 AD/NAD/ACS + 合併對照 HC=NAD∪ACS）。"""
    g = {b: group_of(b) for b in age}
    return [
        ("AD",  {b for b, gg in g.items() if gg == "P"}),
        ("NAD", {b for b, gg in g.items() if gg == "NAD"}),
        ("ACS", {b for b, gg in g.items() if gg == "ACS"}),
        ("HC",  {b for b, gg in g.items() if gg in ("NAD", "ACS")}),
    ]

# ── 表 + 散點（每模型）──────────────────────────────────────────────────────

def build_ancova_blocks(methods_scores, age, comparisons, one_sided=False):
    """每 method × comparison 一列，回 [(label, rows)]；row=(comp, n, unadjΔ, group, age, fit, β3)。

    one_sided=True 時未校正 Δ 與 group β2 改單尾（H1: arm1>arm2）；age β1、slope-diff β3 恆雙尾。
    """
    blocks = []
    for label, scores in methods_scores:
        rows = []
        for n1, s1, n2, s2 in comparisons:
            fit = _ancova_fit(scores, age, s1, s2)
            if fit is None:
                rows.append((f"{n1} vs {n2}", "NA", "NA", "NA", "NA", None, "NA"))
                continue
            n, unadj, unadj_p, _b0, b1, b2, gp, ap = fit
            if one_sided:
                unadj_p = unadj_p / 2 if unadj >= 0 else 1 - unadj_p / 2
                gp = gp / 2 if b2 >= 0 else 1 - gp / 2
            sd = _slope_diff_fit(scores, age, s1, s2)   # β3 恆雙尾
            b3_s = "NA" if sd is None else f"{_coef(sd[0])} ({_pstr(sd[1])})"
            rows.append((f"{n1} vs {n2}", n,
                         f"{_coef(unadj)} ({_pstr(unadj_p)})",
                         f"{_coef(b2)} ({_pstr(gp)})",
                         f"{_coef(b1)} ({_pstr(ap)})", fit, b3_s))
        blocks.append((label, rows))
    return blocks


_COL_W_ANCOVA = {"A": 24.7, "B": 18.0, "C": 12.0, "D": 8.0,
                 "E": 26.0, "F": 26.0, "G": 24.0, "H": 26.0}


def write_ancova_xlsx(model, bg_mode, blocks, out_path, col_w=None):
    """寫 ANCOVA 表：每 method × comparison 一列，合併 Model/Method 欄。col_w 可覆寫欄寬。"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"ancova_{'bg' if bg_mode == 'background' else 'nobg'}"
    ws.append(["Embedding Model", "Asymmetry Method", "Comparison", "n",
               "unadjusted Δ (p)", "age-adj group β (p)", "age β (p)",
               "slope-diff β3 (p)"])
    first = 2
    r = first
    method_spans = []
    for label, rows in blocks:
        bs = r
        for comp, n, unadj_s, grp_s, age_s, _, b3_s in rows:
            ws.cell(r, 3, comp)
            ws.cell(r, 4, n)
            ws.cell(r, 5, unadj_s)
            ws.cell(r, 6, grp_s)
            ws.cell(r, 7, age_s)
            ws.cell(r, 8, b3_s)
            r += 1
        ws.cell(bs, 2, label)
        method_spans.append((bs, r - 1))
    ws.cell(first, 1, MODEL_DISPLAY.get(model, model))
    last = r - 1

    for row in ws.iter_rows(min_row=1, max_row=last, max_col=8):
        for c in row:
            c.alignment = _CENTER
            c.border = _BORDER
            c.font = (_F_HEADER if c.row == 1
                      else _F_BIG if c.column_letter in ("A", "B")
                      else _F_DATA)
    ws.merge_cells(start_row=first, start_column=1, end_row=last, end_column=1)
    for a, b in method_spans:
        ws.merge_cells(start_row=a, start_column=2, end_row=b, end_column=2)
    for col, w in (col_w or _COL_W_ANCOVA).items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = _ROW_H_HEADER
    for rr in range(first, last + 1):
        ws.row_dimensions[rr].height = _ROW_H_DATA
    wb.save(out_path)
    logger.info(f"saved {out_path}")


def write_ancova_scatter(scores, age, ad, hc, fit, title, out_path):
    """score-vs-age 散點 + 兩條平行 ANCOVA 迴歸線（HC 藍 / AD 紅），標註年齡校正後組別 β、p。"""
    if fit is None:
        return
    _, _, _, b0, b1, b2, gp, _ = fit
    ad_pts = [(age[b], scores[b]) for b in ad if b in scores and b in age]
    hc_pts = [(age[b], scores[b]) for b in hc if b in scores and b in age]
    if len(ad_pts) < 2 or len(hc_pts) < 2:
        return
    fig, ax = plt.subplots(figsize=(6, 5))
    for pts, color, lab in [(hc_pts, "#4C72B0", "HC"), (ad_pts, "#C44E52", "AD")]:
        xs, ys = zip(*pts)
        ax.scatter(xs, ys, s=10, alpha=0.35, color=color, label=lab)
    ages = [a for a, _ in ad_pts + hc_pts]
    xr = np.array([min(ages), max(ages)])
    ax.plot(xr, b0 + b1 * xr, color="#4C72B0", lw=2)             # HC 迴歸線
    ax.plot(xr, b0 + b2 + b1 * xr, color="#C44E52", lw=2)        # AD 迴歸線（與 HC 平行）
    sd = _slope_diff_fit(scores, age, ad, hc)
    b3txt = "" if sd is None else f"; slope-diff β3={sd[0]:+.4g}, p={_pstr(sd[1])}"
    ax.set_xlabel("Age")
    ax.set_ylabel("Asymmetry score (per subject)")
    ax.set_title(f"{title}\nage-adj group β={b2:+.4g}, p={_pstr(gp)}{b3txt}")
    ax.legend()
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(str(out_path), dpi=150, bbox_inches="tight")
    plt.close(fig)
    logger.info(f"saved {out_path}")


def run_table(cohort, model, bg_mode, out_base):
    """單模型：ANCOVA 表（asymmetry_score_stat_ancova.xlsx）+ 每 method 的 AD-vs-HC 散點。"""
    out_base.mkdir(parents=True, exist_ok=True)
    full_df = cohort_list(*cohort)
    methods_scores = [(label, scorer_scores(full_df, model, variant, bg_mode, method))
                      for label, variant, method, _ in METHODS]
    full_comps = full_comparisons(full_df)
    age = per_subject_age(full_df)
    ancova_blocks = build_ancova_blocks(methods_scores, age, full_comps)
    write_ancova_xlsx(model, bg_mode, ancova_blocks,
                      out_base / "asymmetry_score_stat_ancova.xlsx")
    ad_set, hc_set = full_comps[0][1], full_comps[0][3]               # AD-vs-HC
    for (_lbl, _v, _m, safe), (label, scores), (_, arows) in zip(
            METHODS, methods_scores, ancova_blocks):
        write_ancova_scatter(scores, age, ad_set, hc_set, arows[0][5],
                             f"{label} (AD vs HC)", out_base / "ancova" / f"{safe}.png")

# ── 跨模型 4×8 總圖 ─────────────────────────────────────────────────────────

def build_model_grid(cohort, bg_mode, out_path):
    """四模型 × 八 method 的 ANCOVA 散點總圖（AD-vs-HC，全體 subject、年齡平行斜率擬合）。"""
    df = cohort_list(*cohort)
    age = per_subject_age(df)
    ad = {b for b in age if group_of(b) == "P"}
    hc = {b for b in age if group_of(b) in ("NAD", "ACS")}

    fig, axes = plt.subplots(len(MODELS), len(METHODS),
                             figsize=(3.7 * len(METHODS), 14), sharex=True)
    for i, mkey in enumerate(MODELS):
        for j, (_label, variant, method, _safe) in enumerate(METHODS):
            ax = axes[i][j]
            sc = scorer_scores(df, mkey, variant, bg_mode, method)
            fit = _ancova_fit(sc, age, ad, hc)
            ad_pts = [(age[b], sc[b]) for b in ad if b in sc and b in age]
            hc_pts = [(age[b], sc[b]) for b in hc if b in sc and b in age]
            for pts, color, lab in [(hc_pts, "#4C72B0", "HC"), (ad_pts, "#C44E52", "AD")]:
                if pts:
                    xs, ys = zip(*pts)
                    ax.scatter(xs, ys, s=6, alpha=0.30, color=color, label=lab)
            if fit:
                _, _, _, b0, b1, b2, gp, _ = fit
                ages = [a for a, _ in ad_pts + hc_pts]
                xr = np.array([min(ages), max(ages)])
                ax.plot(xr, b0 + b1 * xr, color="#4C72B0", lw=1.8)           # HC
                ax.plot(xr, b0 + b2 + b1 * xr, color="#C44E52", lw=1.8)      # AD（平行）
                sd = _slope_diff_fit(sc, age, ad, hc)
                b3line = "" if sd is None else f"\nβ₃={sd[0]:+.3g} (p={_pstr(sd[1])})"
                ax.text(0.04, 0.96,
                        f"adj β₂={b2:+.3g} (p={_pstr(gp)}){b3line}",
                        transform=ax.transAxes, va="top", ha="left", fontsize=8,
                        bbox=dict(boxstyle="round", fc="white", ec="0.7", alpha=0.85))
            if i == 0:
                ax.set_title(COL_LABELS[j], fontsize=11, fontweight="bold", pad=8)
            if j == 0:
                ax.set_ylabel(f"{MODEL_DISPLAY.get(mkey, mkey)}\n\nasymmetry score",
                              fontsize=11, fontweight="bold")
            if i == len(MODELS) - 1:
                ax.set_xlabel("Age", fontsize=10)
        logger.info(f"row done: {mkey}")

    axes[0][-1].legend(loc="lower right", fontsize=9, framealpha=0.9)
    fig.suptitle("ANCOVA — asymmetry score vs age (AD vs HC), parallel-slope fit "
                 "(β₂ = age-adj gap; β₃ = age×group slope diff, 2-sided p)",
                 fontsize=15, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, 0.97))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(str(out_path), dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    logger.info(f"saved {out_path}")

# ── ArcFace 32 格 ───────────────────────────────────────────────────────────

def _cell(ax, scores, age, case, ctrl, fit):
    """單格 score-vs-age 散點 + 兩條平行 ANCOVA 線，左上標 β0/β1/β2（單尾）與 β3（雙尾）。"""
    case_pts = [(age[b], scores[b]) for b in case if b in scores and b in age]
    ctrl_pts = [(age[b], scores[b]) for b in ctrl if b in scores and b in age]
    for pts, color in [(ctrl_pts, _CTRL), (case_pts, _CASE)]:
        if pts:
            xs, ys = zip(*pts)
            ax.scatter(xs, ys, s=6, alpha=0.30, color=color)
    if fit is None:
        return
    _, _, _, b0, b1, b2, gp, _ = fit
    ages = [a for a, _ in case_pts + ctrl_pts]
    if not ages:
        return
    xr = np.array([min(ages), max(ages)])
    ax.plot(xr, b0 + b1 * xr, color=_CTRL, lw=1.6)            # control（group=0）
    ax.plot(xr, b0 + b2 + b1 * xr, color=_CASE, lw=1.6)       # case（group=1，+β2）
    sd = _slope_diff_fit(scores, age, case, ctrl)
    b3_line = ("$\\beta_3$: NA" if sd is None
               else f"$\\beta_3$={sd[0]:+.3g} (age×grp), p={_pstr(sd[1])}")
    txt = (f"$\\beta_0$={b0:+.3g}, $\\beta_1$={b1:+.3g} (age)\n"
           f"$\\beta_2$={b2:+.3g} (group), p={_pstr(_one_sided_p(b2, gp))}\n"
           f"{b3_line}\n"
           r"($\beta_2$: 1-sided >0; $\beta_3$: 2-sided)")
    ax.text(0.04, 0.96, txt, transform=ax.transAxes, va="top", ha="left",
            fontsize=7, bbox=dict(boxstyle="round", fc="white", ec="0.7", alpha=0.85))


def build_arcface_grid(cohort, bg_mode, level, caliper, out_path):
    """ArcFace 大網格：列＝(full/matched × 四族群比較)、欄＝四種 norm·asymmetry 組合。"""
    df = cohort_list(*cohort)
    age = per_subject_age(df)
    matched_comps, _ = matched_comparisons(df, cohort, level, caliper)
    slices = [("full", full_comparisons(df)), ("matched", matched_comps)]  # 各 4 個比較
    col_scores = [scorer_scores(df, _ARCFACE, v, bg_mode, m) for _, v, m in ARCFACE_COLS]

    ncols = len(ARCFACE_COLS)
    nrows = sum(len(comps) for _, comps in slices)   # 2 × 4 = 8
    fig, axes = plt.subplots(nrows, ncols, figsize=(3.7 * ncols, 2.8 * nrows), sharex=True)

    row = 0
    for slice_label, comps in slices:
        for n1, s1, n2, s2 in comps:
            for col, (clabel, _v, _m) in enumerate(ARCFACE_COLS):
                ax = axes[row][col]
                _cell(ax, col_scores[col], age, s1, s2,
                      _ancova_fit(col_scores[col], age, s1, s2))
                if row == 0:
                    ax.set_title(clabel, fontsize=12, fontweight="bold", pad=8)
                if col == 0:
                    ax.set_ylabel(f"{slice_label}\n{n1} vs {n2}\n\nscore",
                                  fontsize=10, fontweight="bold")
                if row == nrows - 1:
                    ax.set_xlabel("Age", fontsize=10)
            logger.info(f"row done: {slice_label} | {n1} vs {n2}")
            row += 1

    handles = [Line2D([0], [0], marker="o", color="w", markerfacecolor=_CASE,
                      markersize=8, label="1st group (case, group=1)"),
               Line2D([0], [0], marker="o", color="w", markerfacecolor=_CTRL,
                      markersize=8, label="2nd group (control, group=0)")]
    axes[0][-1].legend(handles=handles, loc="lower right", fontsize=8, framealpha=0.9)
    fig.suptitle(f"{MODEL_DISPLAY[_ARCFACE]} — ANCOVA: asymmetry score vs age, parallel-slope fit "
                 r"($\beta_2$ = age-adjusted case−control gap, 1-sided p ($\beta_2$>0, case>control); "
                 r"$\beta_3$ = age×group slope difference, 2-sided p ($\beta_3\neq$0))",
                 fontsize=13, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, 0.975))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(str(out_path), dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    logger.info(f"saved {out_path}")

# ── ArcFace 切分：單 slice × {L1,L2}·diff（4×2）───────────────────────────────

def _arcface_slice_fig(slice_label, comps, col_scores, age, out_path):
    """單一 slice（full / matched）的 4 族群比較 × {L1,L2}·diff ANCOVA 圖（4 列 × 2 欄）。"""
    cols = ARCFACE_DIFF_COLS
    nrows, ncols = len(comps), len(cols)
    fig, axes = plt.subplots(nrows, ncols, figsize=(3.7 * ncols, 2.8 * nrows),
                             sharex=True, squeeze=False)
    for row, (n1, s1, n2, s2) in enumerate(comps):
        for col, (clabel, _v, _m) in enumerate(cols):
            ax = axes[row][col]
            _cell(ax, col_scores[col], age, s1, s2,
                  _ancova_fit(col_scores[col], age, s1, s2))
            if row == 0:
                ax.set_title(clabel, fontsize=12, fontweight="bold", pad=8)
            if col == 0:
                ax.set_ylabel(f"{n1} vs {n2}\n\nscore", fontsize=10, fontweight="bold")
            if row == nrows - 1:
                ax.set_xlabel("Age", fontsize=10)
        logger.info(f"row done: {slice_label} | {n1} vs {n2}")

    handles = [Line2D([0], [0], marker="o", color="w", markerfacecolor=_CASE,
                      markersize=8, label="1st group (case, group=1)"),
               Line2D([0], [0], marker="o", color="w", markerfacecolor=_CTRL,
                      markersize=8, label="2nd group (control, group=0)")]
    axes[0][-1].legend(handles=handles, loc="lower right", fontsize=8, framealpha=0.9)
    fig.suptitle(f"{MODEL_DISPLAY[_ARCFACE]} — ANCOVA ({slice_label}, diff only): "
                 r"asymmetry score vs age, parallel-slope fit "
                 r"($\beta_2$ = age-adj gap [1-sided p>0]; $\beta_3$ = age×group slope diff "
                 r"[2-sided p$\neq$0])",
                 fontsize=12, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, 0.96))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(str(out_path), dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    logger.info(f"saved {out_path}")


def build_arcface_diff_splits(cohort, bg_mode, level, caliper, out_dir):
    """由 32 格圖切出 full / matched 兩張 4×2（只 L1·diff、L2·diff）。"""
    df = cohort_list(*cohort)
    age = per_subject_age(df)
    matched_comps, _ = matched_comparisons(df, cohort, level, caliper)
    col_scores = [scorer_scores(df, _ARCFACE, v, bg_mode, m) for _, v, m in ARCFACE_DIFF_COLS]
    _arcface_slice_fig("full", full_comparisons(df), col_scores, age,
                       out_dir / "ancova_grid_arcface_full.png")
    _arcface_slice_fig("matched", matched_comps, col_scores, age,
                       out_dir / "ancova_grid_arcface_matched.png")

# ── per-group 年齡斜率（每組各建一條 score~age）───────────────────────────────

def build_age_slope_table(cohort, model, bg_mode, out_path):
    """每組各自迴歸 score~β0+β1·age 的表：列＝8 method，欄＝AD/NAD/ACS/HC 的 β1（單尾 p，H1: β1>0）。"""
    df = cohort_list(*cohort)
    age = per_subject_age(df)
    groups = _age_slope_groups(age)
    methods_scores = [(label, scorer_scores(df, model, variant, bg_mode, method))
                      for label, variant, method, _ in METHODS]

    wb = Workbook()
    ws = wb.active
    ws.title = f"age_slope_{'bg' if bg_mode == 'background' else 'nobg'}"
    ws.append(["Embedding Model", "Asymmetry Method"]
              + [f"{g} β1 (p) [n={len(s)}]" for g, s in groups])
    first = 2
    r = first
    for label, scores in methods_scores:
        ws.cell(r, 2, label)
        for j, (_g, gset) in enumerate(groups):
            sf = _slope_fit(scores, age, gset)
            if sf is None:
                cell = "NA"
            else:
                _n, _b0, b1, b1_p2 = sf
                cell = f"{_coef(b1)} ({_pstr(_one_sided_p(b1, b1_p2))})"
            ws.cell(r, 3 + j, cell)
        r += 1
    ws.cell(first, 1, MODEL_DISPLAY.get(model, model))
    last = r - 1
    ncol = 2 + len(groups)

    for row in ws.iter_rows(min_row=1, max_row=last, max_col=ncol):
        for c in row:
            c.alignment = _CENTER
            c.border = _BORDER
            c.font = (_F_HEADER if c.row == 1
                      else _F_BIG if c.column_letter in ("A", "B") else _F_DATA)
    ws.merge_cells(start_row=first, start_column=1, end_row=last, end_column=1)
    ws.column_dimensions["A"].width = 24.7
    ws.column_dimensions["B"].width = 18.0
    for j in range(len(groups)):
        ws.column_dimensions[chr(ord("C") + j)].width = 22.0
    ws.row_dimensions[1].height = _ROW_H_HEADER
    for rr in range(first, last + 1):
        ws.row_dimensions[rr].height = _ROW_H_DATA
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info(f"saved {out_path}")


def build_age_slope_fig(cohort, model, bg_mode, out_path):
    """列＝4 組（AD/NAD/ACS/HC）、欄＝L1·diff、L2·diff；每格單組 score~age 散點+迴歸線，標 β1 單尾 p。"""
    df = cohort_list(*cohort)
    age = per_subject_age(df)
    groups = _age_slope_groups(age)
    col_scores = [scorer_scores(df, model, v, bg_mode, m) for _, v, m in ARCFACE_DIFF_COLS]

    nrows, ncols = len(groups), len(ARCFACE_DIFF_COLS)
    fig, axes = plt.subplots(nrows, ncols, figsize=(4.2 * ncols, 2.8 * nrows),
                             sharex=True, squeeze=False)
    for row, (gname, gset) in enumerate(groups):
        color = _GROUP_COLOR[gname]
        for col, (clabel, _v, _m) in enumerate(ARCFACE_DIFF_COLS):
            ax = axes[row][col]
            sc = col_scores[col]
            pts = [(age[b], sc[b]) for b in gset if b in sc and b in age]
            if pts:
                xs, ys = zip(*pts)
                ax.scatter(xs, ys, s=6, alpha=0.30, color=color)
            sf = _slope_fit(sc, age, gset)
            if sf and pts:
                _n, b0, b1, b1_p2 = sf
                xr = np.array([min(xs), max(xs)])
                ax.plot(xr, b0 + b1 * xr, color=color, lw=1.9)
                ax.text(0.04, 0.96,
                        f"$\\beta_1$={b1:+.3g} (age)\n"
                        f"p={_pstr(_one_sided_p(b1, b1_p2))} (1-sided, $\\beta_1$>0)\n"
                        f"n={_n}",
                        transform=ax.transAxes, va="top", ha="left", fontsize=7,
                        bbox=dict(boxstyle="round", fc="white", ec="0.7", alpha=0.85))
            if row == 0:
                ax.set_title(clabel, fontsize=12, fontweight="bold", pad=8)
            if col == 0:
                ax.set_ylabel(f"{gname}\n\nasymmetry score", fontsize=11, fontweight="bold")
            if row == nrows - 1:
                ax.set_xlabel("Age", fontsize=10)
        logger.info(f"row done: {gname}")

    fig.suptitle(f"{MODEL_DISPLAY.get(model, model)} — per-group age slope: score ~ β0 + β1·age "
                 r"(1-sided p tests $\beta_1$>0, i.e. asymmetry rises with age)",
                 fontsize=13, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, 0.97))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(str(out_path), dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    logger.info(f"saved {out_path}")

# ── 主流程 ──────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--mode", choices=["table", "grid", "arcface", "age-slope"], default="table",
                    help="table=單模型表+散點；grid=4×8 跨模型總圖；arcface=ArcFace 32 格；"
                         "age-slope=每組各建 score~age 迴歸（AD/NAD/ACS/HC 的 β1 單尾檢定）")
    ap.add_argument("--p-visit", choices=list(P_VISIT_TOKENS), default=DEFAULT_TOKENS[0])
    ap.add_argument("--p-score", choices=list(P_SCORE_TOKENS), default=DEFAULT_TOKENS[1])
    ap.add_argument("--hc-visit", choices=list(HC_VISIT_TOKENS), default=DEFAULT_TOKENS[2])
    ap.add_argument("--hc-score", choices=list(HC_SCORE_TOKENS), default=DEFAULT_TOKENS[3])
    ap.add_argument("--model", default="arcface", help="僅 --mode table 用")
    ap.add_argument("--bg-mode", choices=["background", "no_background"], default="background")
    ap.add_argument("--match-level", choices=["subject", "visit"], default="visit",
                    help="僅 --mode arcface 用（matched 切片的配對層級）")
    ap.add_argument("--caliper", type=float, default=1.0, help="僅 --mode arcface 用")
    ap.add_argument("--output-dir", type=Path, default=None,
                    help="僅 --mode table；覆寫輸出目錄，留空依 cohort 自動決定")
    args = ap.parse_args()

    cohort = (args.p_visit, args.p_score, args.hc_visit, args.hc_score)
    base = EMBEDDING_FEATURE_STAT_DIR / cohort_path(*cohort) / args.bg_mode
    logger.info(f"cohort = {cohort}  bg = {args.bg_mode}  mode = {args.mode}")

    if args.mode == "table":
        out_base = args.output_dir or (base / args.model)
        run_table(cohort, args.model, args.bg_mode, out_base)
    elif args.mode == "age-slope":
        out_base = args.output_dir or (base / args.model)
        out_base.mkdir(parents=True, exist_ok=True)
        build_age_slope_table(cohort, args.model, args.bg_mode,
                              out_base / "asymmetry_age_slope.xlsx")
        build_age_slope_fig(cohort, args.model, args.bg_mode,
                            out_base / "asymmetry_age_slope.png")
    elif args.mode == "grid":
        build_model_grid(cohort, args.bg_mode, base / "ancova_grid.png")
    else:  # arcface
        ancova_dir = base / _ARCFACE / "ancova"
        build_arcface_grid(cohort, args.bg_mode, args.match_level, args.caliper,
                           ancova_dir / "ancova_grid_arcface.png")
        build_arcface_diff_splits(cohort, args.bg_mode, args.match_level, args.caliper,
                                  ancova_dir)


if __name__ == "__main__":
    main()
