"""鏡射不對稱的族群描述統計（四種 scorer）+ 小提琴圖。

本檔同時是 asymmetry 分析的「共用底層」：scorer 分數、比較集合、樣式常數都定義在此，
ANCOVA 相關（表 / 散點 / 跨模型總圖）拆到 asymmetry_ancova.py，從這裡 import 共用函式。

每 subject 一個分數，四種 scorer（皆零重新提取、共用已落地 .npy；全走
src.embedding.classification.build_scorer）：
  Σ|l − r|        —— 差異向量 L1 norm（l1_norm；純 norm、無 CV）
  √Σ(l − r)²      —— 差異向量 L2 norm（l2_norm；純 norm、無 CV）
  centroid_dist   —— 離 AD/HC 質心的 cosine 距離差（監督，fit AD-vs-HC，GroupKFold OOF）
  lda_projection  —— LDA 判別軸投影（監督，fit AD-vs-HC，GroupKFold OOF）
一律 per-pair（photo_mode="all"）：每個 pair 各自評分 → _pool_to_id 聚到 visit → 跨 visit 平均
成每 subject 一分；GroupKFold 依 subject 分組防 visit 洩漏。監督式分數越高＝越像 AD，配對前後
共用同一份 forward OOF、差別只在納入哪些 subject。

主表（xlsx）每個 method × 4 組比較 AD-vs-{HC,NAD,ACS} 與 NAD-vs-ACS，含 mean±SD、
Welch t-test p、以及「score 直接當預測」的 ROC AUC（arm1 為正類）。輸出兩個檔：
  asymmetry_score_stat.xlsx              —— full（未配對；NAD-vs-ACS = 全體直接比）
  asymmetry_score_stat_1by1matched.xlsx  —— 1:1 配對（AD-vs-HC 一條全域配對切片；
                                            NAD-vs-ACS 另做一條 NAD-vs-ACS 年齡配對）
另出 violin/{full,1by1matched}/<method>.png（三族群分布）。
"""

import argparse
import logging
import sys
from collections import defaultdict
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from scipy import stats as sp_stats
from sklearn.metrics import roc_auc_score

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))  # scripts/
from _paths import PROJECT_ROOT  # noqa: F401

from src.config import (
    EMBEDDING_FEATURE_STAT_DIR,
    cohort_path,
    P_VISIT_TOKENS, P_SCORE_TOKENS, HC_VISIT_TOKENS, HC_SCORE_TOKENS,
)
from src.common.cohort import cohort_list, base_id_of, group_of
from src.common.features import load_feature_matrix
from src.common.matching import match_by_age, split_by_group, age_match
from src.embedding.classification import build_scorer, train

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# 對齊既有 embedding 分類常用的 cohort（與 repo DEFAULT_COHORT_TOKENS 不同：HC 取 all、P 取 cdrall）
DEFAULT_TOKENS = ("p_first", "p_cdrall", "hc_all", "hc_cdrall_or_mmseall")
# (顯示名, 載入 variant, scorer 名, violin 檔名)。順序＝表中由上到下的 method 區塊。
# 乾淨版：2 種 asymmetry vector（diff / rel_diff，皆帶號）× 4 種 scorer = 8 區塊。
# abs 不另列為 variant：norm 對 sign 不敏感（l1/l2 吃帶號＝吃 |·|）；centroid/lda 用帶號 diff
# （|diff| 全非負會使兩質心近共線、centroid 退化成 chance，帶號才分得開）。四者皆走 build_scorer。
METHODS = [
    ("L1 · diff", "differences", "l1_norm", "l1_diff"),
    ("L2 · diff", "differences", "l2_norm", "l2_diff"),
    ("CentroidDist · diff", "differences", "centroid_dist", "cd_diff"),
    ("LDA · diff", "differences", "lda_projection", "lda_diff"),
    ("L1 · rel_diff", "relative_differences", "l1_norm", "l1_rel"),
    ("L2 · rel_diff", "relative_differences", "l2_norm", "l2_rel"),
    ("CentroidDist · rel_diff", "relative_differences", "centroid_dist", "cd_rel"),
    ("LDA · rel_diff", "relative_differences", "lda_projection", "lda_rel"),
]
_SCORER_PHOTO_MODE = "all"  # scorer 的 fit/OOF 表徵（per-pair，對齊既有分類組合）
GROUPS = ["ACS", "NAD", "P"]  # violin
MODEL_DISPLAY = {"arcface": "ArcFace", "topofr": "TopoFR", "dlib": "dlib", "vggface": "VGGFace"}

# ── 分數 ────────────────────────────────────────────────────────────────

def scorer_scores(cohort_df, model, variant, bg_mode, method) -> dict:
    """某 scorer 的每 subject 分數：forward GroupKFold OOF（依 subject 分組防洩漏），跨 visit 平均。

    對 per-pair 向量(photo_mode="all")套 build_scorer(method)：l1_norm/l2_norm 直接取 norm(無 CV)，
    centroid_dist/lda_projection fit AD(P=1)-vs-HC(else=0) 取 OOF decision_function(越高越像 AD)。
    _pool_to_id 先把每列(pair)分數聚到 visit、再於此跨 visit 平均成每 subject 一個分數。

    Returns:
        dict base_id -> score。缺 .npy 者由 load_feature_matrix 略過；全缺的 subject 不入列。
    """
    label_map = {i: int(g == "P") for i, g in zip(cohort_df["ID"], cohort_df["Group"])}
    X, row_ids = load_feature_matrix(
        cohort_df["ID"].tolist(), model, variant, bg_mode, _SCORER_PHOTO_MODE)
    if len(X) == 0:
        return {}
    y = np.array([label_map[i] for i in row_ids], dtype=int)
    _, score_method, needs_cv = build_scorer(method)
    oof = train(X, row_ids, y, lambda: build_scorer(method)[0],
                score_method, needs_cv, "forward")
    per_subj = defaultdict(list)
    for sid, sc in zip(oof["ID"], oof["y_score"]):
        per_subj[base_id_of(sid)].append(float(sc))
    return {b: float(np.mean(v)) for b, v in per_subj.items()}

# ── 比較的受試者集合（與 method 無關，只看 cohort / 配對）──────────────────

def _grp_subjects(cohort_df):
    g = defaultdict(set)
    for i in cohort_df["ID"]:
        g[group_of(base_id_of(i))].add(base_id_of(i))
    return g


def full_comparisons(cohort_df):
    """未配對：每組全體 subject。回 [(name1, set1, name2, set2)]。"""
    g = _grp_subjects(cohort_df)
    return [("AD", g["P"], "HC", g["NAD"] | g["ACS"]),
            ("AD", g["P"], "NAD", g["NAD"]),
            ("AD", g["P"], "ACS", g["ACS"]),
            ("NAD", g["NAD"], "ACS", g["ACS"])]


def matched_comparisons(cohort_df, cohort, level, caliper):
    """配對：AD-vs-{HC,NAD,ACS} 取單一 AD-vs-HC 全域配對切片；NAD-vs-ACS 另做一條配對。"""
    p_ids, hc_ids = match_by_age(*cohort, level=level, priority=["ACS"], caliper=caliper)
    pairs = list(zip(p_ids, hc_ids))
    ad = lambda keep: {base_id_of(p) for p, h in pairs if group_of(base_id_of(h)) in keep}
    hc = lambda keep: {base_id_of(h) for p, h in pairs if group_of(base_id_of(h)) in keep}

    nad_ids, acs_ids = age_match(split_by_group(cohort_df, case="NAD", control="ACS"),
                                 level=level, priority=["ACS"], caliper=caliper)
    comps = [("AD", ad({"NAD", "ACS"}), "HC", hc({"NAD", "ACS"})),
             ("AD", ad({"NAD"}), "NAD", hc({"NAD"})),
             ("AD", ad({"ACS"}), "ACS", hc({"ACS"})),
             ("NAD", {base_id_of(x) for x in nad_ids},
              "ACS", {base_id_of(x) for x in acs_ids})]
    ad_hc_subjects = {base_id_of(i) for i in set(p_ids) | set(hc_ids)}  # 供 violin
    return comps, ad_hc_subjects

# ── 主表（xlsx）────────────────────────────────────────────────────────

def _fmt(a):
    return f"{a.mean():.4f}±{a.std(ddof=1):.4f}" if len(a) else "NA"


def _pstr(p):
    star = "**" if p < 0.01 else "*" if p < 0.05 else ""
    return f"{p:.4e}{star}"


def _auc(a1, a2):
    """以 score 直接分類 arm1-vs-arm2 的 ROC AUC（arm1 為正類）。>0.5=arm1 分數偏高。"""
    y = np.r_[np.ones(len(a1)), np.zeros(len(a2))]
    return roc_auc_score(y, np.r_[a1, a2])


def build_blocks(methods_scores, comparisons, one_sided=False):
    """methods_scores: [(label, scores_dict)]。回 [(label, rows)]，rows 每 2 列一組比較。

    one_sided=True 時 p 改為單尾、方向 H1: arm1 > arm2（arm1＝比較的前者/case）：
    由雙尾 Welch p 與兩臂均值方向換算（mean1≥mean2 取 p/2，否則 1−p/2；對 t 分布精確）。
    """
    blocks = []
    for label, scores in methods_scores:
        rows = []
        for n1, s1, n2, s2 in comparisons:
            a1 = np.array([scores[b] for b in s1 if b in scores])
            a2 = np.array([scores[b] for b in s2 if b in scores])
            p = sp_stats.ttest_ind(a1, a2, equal_var=False).pvalue
            if one_sided and len(a1) and len(a2):
                p = p / 2 if a1.mean() >= a2.mean() else 1 - p / 2
            rows.append((n1, len(a1), _fmt(a1), _pstr(p), f"{_auc(a1, a2):.4f}"))
            rows.append((n2, len(a2), _fmt(a2), None, None))
        blocks.append((label, rows))
    return blocks


# 樣式對齊參考表（Times New Roman、四邊 thin 框線、置中）—— ancova 表亦 import 這組常數。
_TNR = "Times New Roman"
_BORDER = Border(*(Side(style="thin"),) * 4)
_CENTER = Alignment(horizontal="center", vertical="center")
_F_HEADER = Font(name=_TNR, size=12, bold=True)   # 表頭
_F_BIG = Font(name=_TNR, size=14, bold=True)      # Model(A) / Method(B)
_F_GROUP = Font(name=_TNR, size=10, bold=True)    # Group(C)
_F_DATA = Font(name=_TNR, size=10, bold=False)    # n / value / p / AUC
_COL_W = {"A": 24.7, "B": 28.0, "C": 12.6, "D": 10.6, "E": 30.6, "F": 18.6, "G": 12.0}
_ROW_H_HEADER, _ROW_H_DATA = 24, 18  # 標題列 / 數值列 列高


def write_xlsx(model, bg_mode, blocks, out_path, col_w=None):
    """blocks: [(label, rows)]，rows 每 2 列一組比較。合併 Model/Method/p/AUC；n 僅兩臂相等時合併。

    col_w：欄寬 dict（預設 _COL_W）；傳入可覆寫（例如 B 欄放數學式時加寬）。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"score_p_value_{'bg' if bg_mode == 'background' else 'nobg'}"
    ws.append(["Embedding Model", "Asymmetry Method", "Group", "n",
               "asymmetry value (mean±SD)", "p-value", "AUC"])

    first = 2
    r = first
    method_spans, pair_spans, n_spans = [], [], []
    for label, rows in blocks:
        bs = r
        for grp, n, val, p, auc in rows:
            ws.cell(r, 3, grp)
            ws.cell(r, 4, n)
            ws.cell(r, 5, val)
            if p is not None:
                ws.cell(r, 6, p)
            if auc is not None:
                ws.cell(r, 7, auc)
            r += 1
        ws.cell(bs, 2, label)
        method_spans.append((bs, r - 1))
        for k in range(len(rows) // 2):
            cs = bs + 2 * k
            pair_spans.append((cs, cs + 1))  # p 與 AUC 每比較跨 2 列合併
            if rows[2 * k][1] == rows[2 * k + 1][1]:  # 兩臂 n 相等才合併 D
                n_spans.append((cs, cs + 1))
    ws.cell(first, 1, MODEL_DISPLAY.get(model, model))
    last = r - 1

    # 先把字型 / 框線 / 對齊套到每一格，再合併（合併區邊框才完整）
    for row in ws.iter_rows(min_row=1, max_row=last, max_col=7):
        for c in row:
            c.alignment = _CENTER
            c.border = _BORDER
            c.font = (_F_HEADER if c.row == 1
                      else _F_BIG if c.column_letter in ("A", "B")
                      else _F_GROUP if c.column_letter == "C"
                      else _F_DATA)

    ws.merge_cells(start_row=first, start_column=1, end_row=last, end_column=1)  # Model
    for a, b in method_spans:
        ws.merge_cells(start_row=a, start_column=2, end_row=b, end_column=2)
    for a, b in n_spans:
        ws.merge_cells(start_row=a, start_column=4, end_row=b, end_column=4)
    for a, b in pair_spans:
        ws.merge_cells(start_row=a, start_column=6, end_row=b, end_column=6)  # p-value
        ws.merge_cells(start_row=a, start_column=7, end_row=b, end_column=7)  # AUC
    for col, w in (col_w or _COL_W).items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = _ROW_H_HEADER            # 標題列
    for rr in range(first, last + 1):
        ws.row_dimensions[rr].height = _ROW_H_DATA         # 數值列
    wb.save(out_path)
    logger.info(f"saved {out_path}")

# ── 小提琴 ──────────────────────────────────────────────────────────────

def write_violin(scores, subjects, title, out_path):
    items = [(g, [scores[b] for b in subjects if b in scores and group_of(b) == g])
             for g in GROUPS]
    items = [(g, d) for g, d in items if len(d) >= 2]  # violinplot 需 >=2 點估 KDE
    if not items:
        logger.warning(f"skip violin（無任一組 >=2 subjects）: {out_path}")
        return
    fig, ax = plt.subplots(figsize=(6, 5))
    parts = ax.violinplot([d for _, d in items], showmedians=True)
    for pc, color in zip(parts["bodies"], ["#4C72B0", "#55A868", "#C44E52"]):
        pc.set_facecolor(color)
        pc.set_alpha(0.6)
    ax.set_xticks(range(1, len(items) + 1))
    ax.set_xticklabels([f"{g}\n(n={len(d)})" for g, d in items])
    ax.set_ylabel("Asymmetry score (per subject)")
    ax.set_title(title)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(str(out_path), dpi=150, bbox_inches="tight")
    plt.close(fig)
    logger.info(f"saved {out_path}")

# ── 主流程 ──────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--p-visit", choices=list(P_VISIT_TOKENS), default=DEFAULT_TOKENS[0])
    ap.add_argument("--p-score", choices=list(P_SCORE_TOKENS), default=DEFAULT_TOKENS[1])
    ap.add_argument("--hc-visit", choices=list(HC_VISIT_TOKENS), default=DEFAULT_TOKENS[2])
    ap.add_argument("--hc-score", choices=list(HC_SCORE_TOKENS), default=DEFAULT_TOKENS[3])
    ap.add_argument("--model", default="arcface")
    ap.add_argument("--bg-mode", choices=["background", "no_background"], default="background")
    ap.add_argument("--match-level", choices=["subject", "visit"], default="visit",
                    help="配對層級；visit=每次拜訪獨立（預設），subject=兩臂等量平衡")
    ap.add_argument("--caliper", type=float, default=1.0)
    ap.add_argument("--output-dir", type=Path, default=None,
                    help="覆寫輸出目錄；留空依 cohort 自動決定")
    args = ap.parse_args()

    cohort = (args.p_visit, args.p_score, args.hc_visit, args.hc_score)
    out_base = args.output_dir or (
        EMBEDDING_FEATURE_STAT_DIR / cohort_path(*cohort) / args.bg_mode / args.model)
    out_base.mkdir(parents=True, exist_ok=True)
    logger.info(f"cohort = {cohort}  match-level = {args.match_level}")
    logger.info(f"output-dir = {out_base}")

    full_df = cohort_list(*cohort)
    methods_scores = [(label, scorer_scores(full_df, args.model, variant, args.bg_mode, method))
                      for label, variant, method, _ in METHODS]

    full_comps = full_comparisons(full_df)
    matched_comps, matched_subjects = matched_comparisons(
        full_df, cohort, args.match_level, args.caliper)

    write_xlsx(args.model, args.bg_mode, build_blocks(methods_scores, full_comps),
               out_base / "asymmetry_score_stat.xlsx")               # full（未配對）
    write_xlsx(args.model, args.bg_mode, build_blocks(methods_scores, matched_comps),
               out_base / "asymmetry_score_stat_1by1matched.xlsx")    # 1:1 配對

    for (_lbl, _v, _m, safe), (label, scores) in zip(METHODS, methods_scores):
        write_violin(scores, set(scores), f"{label} (full)",
                     out_base / "violin" / "full" / f"{safe}.png")
        write_violin(scores, matched_subjects, f"{label} (1by1matched)",
                     out_base / "violin" / "1by1matched" / f"{safe}.png")


if __name__ == "__main__":
    main()
